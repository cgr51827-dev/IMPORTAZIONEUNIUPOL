import io
import zipfile
import hashlib
from copy import copy
from datetime import datetime
from dataclasses import dataclass

import streamlit as st
import openpyxl

st.set_page_config(page_title="Unipol Rental Import Tool", page_icon="📄", layout="wide")

APP_TITLE = "Unipol Rental Import Tool"

IMPORT_STANDARD_MAP = {
    "A": "C",
    "D": "U",
    "E": "T",
    "F": "X",
    "H": "Y",
    "J": "D",
    "K": "E",
    "L": "F",
    "M": "G",
    "U": "H",
}

RECAPITI_CONTACT_OUTPUT_COLUMN = "F"
RECAPITI_CODE_COLUMN = "B"
RECAPITI_SOURCE_CODE_COLUMN = "H"
RECAPITI_SOURCE_CONTACT_COLUMNS = ["Z", "AA", "AB"]

HEADER_SCAN_ROWS = 6
ATTRIBUZIONE_KEYWORD = "ATTRIBUZIONE"


def normalize(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").replace("\r", " ").split()).strip().lower()


def is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def excel_col_to_idx(col: str) -> int:
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - 64)
    return result


def format_value(value):
    if isinstance(value, str):
        return value.strip()
    return value


def copy_row_style(ws, source_row: int, target_row: int):
    if source_row == target_row:
        return
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
    if ws.row_dimensions[source_row].height is not None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def clear_data_area(ws, start_row: int):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def write_records_to_workbook(template_bytes: bytes, records: list[dict]) -> bytes:
    tmp = io.BytesIO(template_bytes)
    wb = openpyxl.load_workbook(tmp)
    ws = wb[wb.sheetnames[0]]

    start_row = 2
    clear_data_area(ws, start_row)
    sample_row = 2 if ws.max_row >= 2 else 1

    for i, record in enumerate(records, start=start_row):
        copy_row_style(ws, sample_row, i)
        for out_col, value in record.items():
            ws[f"{out_col}{i}"] = format_value(value)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_import_standard_records(src_ws):
    records = []
    for row in range(2, src_ws.max_row + 1):
        source_code = src_ws[f"H{row}"].value
        if is_blank(source_code):
            continue
        rec = {}
        for out_col, src_col in IMPORT_STANDARD_MAP.items():
            rec[out_col] = src_ws[f"{src_col}{row}"].value
        records.append(rec)
    return records


def build_recapiti_records(src_ws):
    records = []
    for row in range(2, src_ws.max_row + 1):
        contacts = []
        for src_col in RECAPITI_SOURCE_CONTACT_COLUMNS:
            val = src_ws[f"{src_col}{row}"].value
            if not is_blank(val):
                contacts.append(val)
        if not contacts:
            continue
        code = src_ws[f"{RECAPITI_SOURCE_CODE_COLUMN}{row}"].value
        for contact in contacts:
            rec = {
                RECAPITI_CODE_COLUMN: code,
                RECAPITI_CONTACT_OUTPUT_COLUMN: contact,
            }
            records.append(rec)
    return records


@dataclass
class RateGroup:
    name: str
    start_col: int
    end_col: int
    field_cols: dict


def detect_rate_groups(src_ws):
    max_col = src_ws.max_column
    group_starts = []

    for col in range(excel_col_to_idx("AM"), max_col + 1):
        for r in range(1, min(HEADER_SCAN_ROWS, src_ws.max_row) + 1):
            val = normalize(src_ws.cell(r, col).value)
            if ATTRIBUZIONE_KEYWORD.lower() in val:
                group_starts.append((col, src_ws.cell(r, col).value))
                break

    if not group_starts:
        raise ValueError("Non ho trovato intestazioni 'ATTRIBUZIONE ...' dalla colonna AM in poi nel file sorgente.")

    groups = []
    for i, (start_col, group_name) in enumerate(group_starts):
        end_col = group_starts[i + 1][0] - 1 if i + 1 < len(group_starts) else max_col
        field_cols = {}
        for col in range(start_col, end_col + 1):
            for r in range(1, min(HEADER_SCAN_ROWS, src_ws.max_row) + 1):
                header = normalize(src_ws.cell(r, col).value)
                if "tipo di documento" in header:
                    field_cols["Tipo di documento"] = col
                elif "numero documento" in header:
                    field_cols["Numero documento"] = col
                elif "scadenza al netto" in header:
                    field_cols["Scadenza al netto"] = col
                elif "importo in divisa interna" in header:
                    field_cols["Importo in divisa interna"] = col
        groups.append(RateGroup(str(group_name), start_col, end_col, field_cols))

    if not any(g.field_cols for g in groups):
        raise ValueError(
            "Ho trovato i blocchi ATTRIBUZIONE ma non i campi richiesti: "
            "Tipo di documento / Numero documento / Scadenza al netto / Importo in divisa interna."
        )
    return groups


def build_rate_records(src_ws):
    groups = detect_rate_groups(src_ws)
    records = []
    for row in range(2, src_ws.max_row + 1):
        external_code = src_ws[f"H{row}"].value
        if is_blank(external_code):
            continue
        for group in groups:
            values = {}
            for field_name in [
                "Tipo di documento",
                "Numero documento",
                "Scadenza al netto",
                "Importo in divisa interna",
            ]:
                col = group.field_cols.get(field_name)
                values[field_name] = src_ws.cell(row, col).value if col else None
            if all(is_blank(v) for v in values.values()):
                continue
            rec = {
                "A": external_code,
                "C": values["Tipo di documento"],
                "D": values["Numero documento"],
                "F": values["Scadenza al netto"],
                "G": values["Importo in divisa interna"],
            }
            records.append(rec)
    return records


def sha256_hex(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def check_login():
    st.sidebar.markdown("### Accesso")

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
        st.session_state.username = None

    USERNAME = "RECAP"
    PASSWORD_HASH = sha256_hex("Recap26@")

    if st.session_state.authenticated:
        st.sidebar.success(f"Connesso come: {st.session_state.username}")
        if st.sidebar.button("Esci"):
            st.session_state.authenticated = False
            st.session_state.username = None
            st.rerun()
        return

    with st.sidebar.form("login_form"):
        username = st.text_input("Utente")
        password = st.text_input("Password", type="password")
        submit = st.form_submit_button("Entra")

    if submit:
        if username == USERNAME and sha256_hex(password) == PASSWORD_HASH:
            st.session_state.authenticated = True
            st.session_state.username = username
            st.rerun()
        else:
            st.sidebar.error("Credenziali non valide.")
            st.stop()

    st.info("Inserisci utente e password nella barra laterale.")
    st.stop()


check_login()

st.title(APP_TITLE)
st.caption("Versione web con accesso protetto. Recapiti in colonna F.")

with st.expander("Nota importante", expanded=True):
    st.warning(
        "Questa versione online lavora con template .xlsx. "
        "I due template storici .xls vanno prima convertiti una volta sola in .xlsx."
    )

col1, col2 = st.columns([1.3, 1])

with col1:
    src_file = st.file_uploader("1) Carica il file sorgente PFM Affido", type=["xlsx", "xlsm"])
    tpl_import = st.file_uploader("2) Template Import Standard (.xlsx)", type=["xlsx"], key="tpl_import")
    tpl_recapiti = st.file_uploader("3) Template Recapiti (.xlsx)", type=["xlsx"], key="tpl_recapiti")
    tpl_rate = st.file_uploader("4) Template Rate (.xlsx)", type=["xlsx"], key="tpl_rate")

with col2:
    st.markdown("### Regole attive")
    st.markdown(
        """
- Import Standard con mapping già impostato
- Recapiti: codice in **B**, email/recapito in **F**
- Rate: lettura blocchi da **AM** in poi e trasposizione verticale
        """
    )

go = st.button("Genera file compilati", type="primary", use_container_width=True)

if go:
    if not all([src_file, tpl_import, tpl_recapiti, tpl_rate]):
        st.error("Carica tutti i file richiesti prima di continuare.")
        st.stop()

    try:
        src_wb = openpyxl.load_workbook(io.BytesIO(src_file.getvalue()), data_only=True)
        src_ws = src_wb[src_wb.sheetnames[0]]

        import_records = build_import_standard_records(src_ws)
        recapiti_records = build_recapiti_records(src_ws)
        rate_records = build_rate_records(src_ws)

        out_import = write_records_to_workbook(tpl_import.getvalue(), import_records)
        out_recapiti = write_records_to_workbook(tpl_recapiti.getvalue(), recapiti_records)
        out_rate = write_records_to_workbook(tpl_rate.getvalue(), rate_records)

        zip_buffer = io.BytesIO()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("GeCO_import_standard_compilato.xlsx", out_import)
            zf.writestr("GeCO_recapiti_compilato.xlsx", out_recapiti)
            zf.writestr("GeCO_rate_compilato.xlsx", out_rate)

        zip_buffer.seek(0)

        st.success("Elaborazione completata.")
        c1, c2, c3 = st.columns(3)
        c1.metric("Import Standard", len(import_records))
        c2.metric("Recapiti", len(recapiti_records))
        c3.metric("Rate", len(rate_records))

        st.download_button(
            "Scarica ZIP risultati",
            data=zip_buffer.getvalue(),
            file_name=f"output_geco_{stamp}.zip",
            mime="application/zip",
            use_container_width=True,
        )

    except Exception as e:
        st.exception(e)
